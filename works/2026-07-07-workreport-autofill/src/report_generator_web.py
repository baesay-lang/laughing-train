"""
Excel 보고서 생성 모듈 - 웹 버전 (openpyxl only, win32com 없음)
"""
import re
from io import BytesIO
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.worksheet.properties import PageSetupProperties


# ── 상수 ──────────────────────────────────────────────────────────────
C_TOTAL_BG  = "FFFFFFCC"
TOTAL_SLOTS = 77
FONT_NAME   = "맑은 고딕"


# ── 스타일 헬퍼 ───────────────────────────────────────────────────────
def _side(style: str) -> Side:
    return Side(border_style=style, color="FF000000")


def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


def _align(h="center", v="center", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border_row(col: int, top=None, bottom=None) -> Border:
    def s(style):
        return _side(style) if style else Side()

    left_map  = {1: "medium", 2: "hair", 3: "hair", 4: "hair",
                 5: "hair",   6: None,   7: "hair", 8: "medium"}
    right_map = {1: "hair",   2: "hair", 3: "hair", 4: "hair",
                 5: "hair",   6: "hair", 7: "medium", 8: "medium"}

    return Border(
        left   = s(left_map.get(col)),
        right  = s(right_map.get(col)),
        top    = s(top),
        bottom = s(bottom),
    )


# ── 제목 정제 ─────────────────────────────────────────────────────────
def _strip_title(title: str, code: str) -> str:
    s = title.strip()
    _prefix = re.compile(r'^\[(?:KPI|Non-KPI)\]\s*', re.IGNORECASE)
    # 태그·코드가 어떤 순서로 붙어도 처리되도록 2회 반복
    for _ in range(2):
        while _prefix.match(s):
            s = _prefix.sub('', s).strip()
        if code and code != "미분류":
            s = re.sub(
                r'^' + re.escape(code) + r'(?:-[A-Za-z0-9]+)*[\s:._-]*',
                '', s
            ).strip()
        s = re.sub(
            r'^(?:[A-Z]{1,3}\d{2}(?:-[A-Za-z0-9]+)+'
            r'|[A-Z]-\d+(?:-[A-Za-z0-9]+)+)[\s:._-]+',
            '', s
        ).strip()
    return s or title.strip()


# ── 이벤트 집계 ───────────────────────────────────────────────────────
def _group_events(events: list) -> dict:
    buckets = defaultdict(lambda: {"hours": 0.0, "titles": [], "title_hours": {}})
    for e in events:
        if not e.get("included", True):
            continue
        key = (
            e.get("code",      "미분류"),
            e.get("work_name", e.get("title", "")),
            e.get("code_type", "미분류"),
        )
        h     = e.get("duration_hours", 0.0)
        title = e.get("title", "").strip()
        buckets[key]["hours"] += h
        if title:
            buckets[key]["title_hours"][title] = (
                buckets[key]["title_hours"].get(title, 0.0) + h
            )
            if title not in buckets[key]["titles"]:
                buckets[key]["titles"].append(title)
    return dict(buckets)


def _build_c_text(code_type: str, work_name: str, code: str,
                  all_titles: list, combined_th: dict) -> str:
    b_text = f"[{code_type}] {work_name}"
    seen: dict = {}
    for t in all_titles:
        st = _strip_title(t, code)
        if not st:
            continue
        h = combined_th.get(t, 0.0)
        seen[st] = seen.get(st, 0.0) + h
    if not seen:
        return b_text
    items = list(seen.items())
    if len(items) == 1:
        return f"[{code_type}] {items[0][0]}"
    lines = [f"{i + 1}. {st} ({h:.1f}H)" for i, (st, h) in enumerate(items)]
    return f"[{code_type}]\n" + "\n".join(lines)


def _row_height(c_text: str) -> float:
    """C열 줄 수에 맞춰 행 높이 산정 (기본 80.25, 넘치면 확장)."""
    disp = 0
    for line in c_text.split("\n"):
        disp += max(1, -(-len(line) // 38))
    return max(80.25, disp * 15 + 14)


def _sort_key(k):
    code, work_name, code_type = k
    order = {"KPI": 0, "Non-KPI": 1}
    return (order.get(code_type, 2), code, work_name)


# ── 시트 빌더 ─────────────────────────────────────────────────────────
def _build_sheet(ws, this_events, next_events, user_name, emp_id):

    for col, w in [("A", 13.0), ("B", 38.875), ("C", 46.75),
                   ("D", 17.5),  ("E", 11.75),  ("F", 17.5),
                   ("G", 11.75), ("H", 95.25)]:
        ws.column_dimensions[col].width = w

    this_grp  = _group_events(this_events)
    next_grp  = _group_events(next_events)
    all_keys  = sorted(
        [k for k in set(this_grp) | set(next_grp) if k[2] != "미분류"],
        key=_sort_key,
    )
    this_total = sum(v["hours"] for k, v in this_grp.items() if k[2] != "미분류")
    next_total = sum(v["hours"] for k, v in next_grp.items() if k[2] != "미분류")

    # 행 1: 작성자
    ws.row_dimensions[1].height = 50.25
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value     = f"작성자 : {emp_id} ( {user_name} )"
    c.font      = Font(name=FONT_NAME, size=14, bold=True)
    c.alignment = _align(h="left")
    c.border    = Border(left=_side("thin"), bottom=_side("medium"))
    ws["B1"].border = Border(bottom=_side("medium"))

    # 행 2: 헤더
    ws.row_dimensions[2].height = 25.5
    _hdr = [
        (1, "구분"), (2, "당월 MileStone"), (3, "수행한 업무 상세"),
        (4, "금주(ACTUAL)"), (5, "비중"),
        (6, "차주(PLAN)"),   (7, "비중"),
        (8, "보직자 FEEDBACK"),
    ]
    hdr_fill = _fill("D9D9D9")
    for col_idx, label in _hdr:
        c = ws.cell(row=2, column=col_idx)
        c.value     = label
        c.font      = Font(name=FONT_NAME, size=11, bold=True)
        c.fill      = hdr_fill
        c.alignment = _align()
        c.border    = _border_row(col_idx, top="medium")
        if col_idx in (5, 7):
            c.number_format = "0%"

    # 행 3: 합계
    ws.row_dimensions[3].height = 25.5
    yf = _fill(C_TOTAL_BG)

    def _tc(col_idx, value, fmt=None):
        c = ws.cell(row=3, column=col_idx)
        c.value     = value
        c.font      = Font(name=FONT_NAME, size=11, bold=True)
        c.alignment = _align()
        c.fill      = yf
        c.border    = _border_row(col_idx, top="thin", bottom="thin")
        if fmt:
            c.number_format = fmt

    _tc(1, "총계"); _tc(2, None); _tc(3, None)
    _tc(4, this_total, '0.0"H"'); _tc(5, 1, "0%")
    _tc(6, next_total, '0.0"H"'); _tc(7, 1, "0%")
    _tc(8, None)

    # 행 4~80: 데이터
    for slot in range(TOTAL_SLOTS):
        excel_row = 4 + slot
        ws.row_dimensions[excel_row].height = 80.25

        if slot >= len(all_keys):
            continue

        key  = all_keys[slot]
        code, work_name, code_type = key
        this_h = this_grp.get(key, {}).get("hours", 0.0)
        next_h = next_grp.get(key, {}).get("hours", 0.0)
        this_r = this_h / this_total if this_total > 0 else 0.0
        next_r = next_h / next_total if next_total > 0 else 0.0

        all_titles = []
        combined_th: dict = {}
        for grp in (this_grp.get(key, {}), next_grp.get(key, {})):
            for t in grp.get("titles", []):
                if t and t not in all_titles:
                    all_titles.append(t)
            for t, h in grp.get("title_hours", {}).items():
                combined_th[t] = combined_th.get(t, 0.0) + h

        b_text = f"[{code_type}] {work_name}"
        c_text = _build_c_text(code_type, work_name, code,
                               all_titles, combined_th)

        # C열 항목 수에 맞춰 행 높이 확장
        ws.row_dimensions[excel_row].height = _row_height(c_text)

        top_style = "hair" if slot > 0 else None

        def _dc(col_idx, val, sz=14, h="center", fmt=None, wrap=True,
                _row=excel_row, _top=top_style):
            c = ws.cell(row=_row, column=col_idx)
            c.value     = val
            c.font      = Font(name=FONT_NAME, size=sz, bold=True)
            c.alignment = Alignment(horizontal=h, vertical="center", wrap_text=wrap)
            c.border    = _border_row(col_idx, top=_top, bottom="hair")
            if fmt:
                c.number_format = fmt

        _dc(1, code_type, sz=14)
        c = ws.cell(row=excel_row, column=2)
        c.value     = b_text
        c.font      = Font(name=FONT_NAME, size=14, bold=True)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border    = _border_row(2, top=top_style, bottom="hair")
        _dc(3, c_text, sz=11, h="left")
        _dc(4, this_h, fmt='0.0"H"')
        _dc(5, this_r, fmt="0%")
        _dc(6, next_h, fmt='0.0"H"')
        _dc(7, next_r, fmt="0%")
        _dc(8, "", sz=11)

    # DataBar CF
    last_row = 3 + TOTAL_SLOTS
    for col_letter in ("E", "G"):
        rule = DataBarRule(
            start_type="num", start_value=0,
            end_type="num",   end_value=1,
            color="638EC6",   showValue=True,
        )
        ws.conditional_formatting.add(f"{col_letter}4:{col_letter}{last_row}", rule)

    # 빈 행 숨김 + 동적 인쇄 영역
    n_data = len(all_keys)
    last_data_row = 3 + n_data if n_data > 0 else 3
    for r in range(4 + n_data, 4 + TOTAL_SLOTS):
        ws.row_dimensions[r].hidden = True

    ws.print_area = f"A1:H{last_data_row}"
    ws.print_title_rows = "1:3"   # 2페이지부터 헤더 반복
    ws.page_setup.paperSize   = 9
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.freeze_panes = "A4"
    ws.sheet_view.view = "pageBreakPreview"
    ws.sheet_view.zoomScaleSheetLayoutView = 75


# ── 공개 API ──────────────────────────────────────────────────────────
def generate_report_bytes(
    this_week_events: list,
    next_week_events: list,
    user_name:        str,
    employee_id:      str,
    sheet_name:       str = "",
) -> bytes:
    """
    Excel 보고서를 생성해 bytes로 반환한다.
    Streamlit download_button에 직접 전달 가능.
    """
    if not sheet_name:
        sheet_name = datetime.today().strftime("%y%m%d")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]   # Excel 시트명 최대 31자

    _build_sheet(ws, this_week_events, next_week_events, user_name, employee_id)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
